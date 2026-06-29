"""
Genera un archivo Excel de presupuesto con formato profesional.
"""
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side,
                              numbers)
from openpyxl.utils import get_column_letter


def generar(
    items: list[dict],
    markup_pct: float,
    empresa: str,
    cliente: str,
    numero: int,
    notas: str = "",
    destino: Path | None = None,
    iva_modo: str = "Sin IVA",
) -> Path:
    """
    items: lista de dicts con keys: nombre, sku, precio_costo, cantidad
    markup_pct: porcentaje de ganancia (ej: 30 → +30%)
    Retorna el path del archivo generado.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"

    # Paleta de colores
    AZUL    = "1E3A5F"
    AZUL_C  = "D6E4F0"
    GRIS    = "F5F5F5"
    BLANCO  = "FFFFFF"
    VERDE   = "1A7A4A"

    # Anchos de columna
    ws.column_dimensions["A"].width = 8    # Cant
    ws.column_dimensions["B"].width = 12   # SKU
    ws.column_dimensions["C"].width = 44   # Descripción
    ws.column_dimensions["D"].width = 16   # Precio unit
    ws.column_dimensions["E"].width = 16   # Subtotal

    # ── ENCABEZADO ──────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = empresa.upper()
    c.font  = Font(bold=True, size=18, color=BLANCO)
    c.fill  = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = "PRESUPUESTO MAYORISTA"
    c.font  = Font(bold=True, size=11, color=AZUL)
    c.fill  = PatternFill("solid", fgColor=AZUL_C)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # ── META ───────────────────────────────────────────────────
    fecha = datetime.now().strftime("%d/%m/%Y")
    ws.merge_cells("A3:C3")
    ws["A3"] = f"Cliente: {cliente}"
    ws["A3"].font = Font(bold=True, size=11)

    ws["D3"] = f"N° {numero:05d}"
    ws["D3"].font = Font(bold=True, size=11, color=AZUL)
    ws["D3"].alignment = Alignment(horizontal="right")

    ws["E3"] = fecha
    ws["E3"].alignment = Alignment(horizontal="right")
    ws.row_dimensions[3].height = 20

    ws.row_dimensions[4].height = 6  # espacio

    # ── CABECERA DE TABLA ───────────────────────────────────────
    headers = ["Cant.", "SKU", "Descripción", "Precio unit.", "Subtotal"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font      = Font(bold=True, size=10, color=BLANCO)
        cell.fill      = PatternFill("solid", fgColor=AZUL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
    ws.row_dimensions[5].height = 22

    # ── FILAS DE PRODUCTOS ──────────────────────────────────────
    row = 6
    for i, item in enumerate(items):
        mkp          = item.get("markup", markup_pct)
        precio_venta = item.get("precio_custom") or item["precio_costo"] * (1 + mkp / 100)
        subtotal     = precio_venta * item["cantidad"]
        fondo        = BLANCO if i % 2 == 0 else GRIS

        vals = [
            item["cantidad"],
            item.get("sku") or "",
            item.get("nombre_custom") or item["nombre"],
            precio_venta,
            subtotal,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill   = PatternFill("solid", fgColor=fondo)
            cell.border = _thin_border(light=True)
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 3))

            if col in (4, 5):
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row].height = 18
        row += 1

    # ── TOTAL / IVA ─────────────────────────────────────────────
    subtotal_row = row - 1  # última fila de productos
    row += 1

    ROJO_XL  = "CC1515"
    AMARI_XL = "F5C800"

    if iva_modo == "+ IVA 21%":
        # Fila subtotal
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "SUBTOTAL"
        ws[f"A{row}"].font      = Font(bold=True, size=11, color="333333")
        ws[f"A{row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"E{row}"] = f"=SUM(E6:E{subtotal_row})"
        ws[f"E{row}"].number_format = '"$"#,##0.00'
        ws[f"E{row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 20
        subtotal_cell = f"E{row}"
        row += 1
        # Fila IVA
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "IVA 21%"
        ws[f"A{row}"].font      = Font(bold=True, size=11, color="333333")
        ws[f"A{row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"E{row}"] = f"={subtotal_cell}*0.21"
        ws[f"E{row}"].number_format = '"$"#,##0.00'
        ws[f"E{row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 20
        iva_cell = f"E{row}"
        row += 1
        # Fila total
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "TOTAL A COBRAR"
        ws[f"A{row}"].font      = Font(bold=True, size=12, color=BLANCO)
        ws[f"A{row}"].fill      = PatternFill("solid", fgColor=ROJO_XL)
        ws[f"A{row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"E{row}"] = f"={subtotal_cell}+{iva_cell}"
        ws[f"E{row}"].font          = Font(bold=True, size=12, color=BLANCO)
        ws[f"E{row}"].fill          = PatternFill("solid", fgColor=ROJO_XL)
        ws[f"E{row}"].number_format = '"$"#,##0.00'
        ws[f"E{row}"].alignment     = Alignment(horizontal="right", vertical="center")
    elif iva_modo == "IVA incluido":
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "TOTAL A COBRAR (IVA incl.)"
        ws[f"A{row}"].font      = Font(bold=True, size=12, color=BLANCO)
        ws[f"A{row}"].fill      = PatternFill("solid", fgColor=ROJO_XL)
        ws[f"A{row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"E{row}"] = f"=SUM(E6:E{subtotal_row})"
        ws[f"E{row}"].font          = Font(bold=True, size=12, color=BLANCO)
        ws[f"E{row}"].fill          = PatternFill("solid", fgColor=ROJO_XL)
        ws[f"E{row}"].number_format = '"$"#,##0.00'
        ws[f"E{row}"].alignment     = Alignment(horizontal="right", vertical="center")
    else:  # Sin IVA
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "TOTAL A COBRAR"
        ws[f"A{row}"].font      = Font(bold=True, size=12, color=BLANCO)
        ws[f"A{row}"].fill      = PatternFill("solid", fgColor=ROJO_XL)
        ws[f"A{row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"E{row}"] = f"=SUM(E6:E{subtotal_row})"
        ws[f"E{row}"].font          = Font(bold=True, size=12, color=BLANCO)
        ws[f"E{row}"].fill          = PatternFill("solid", fgColor=ROJO_XL)
        ws[f"E{row}"].number_format = '"$"#,##0.00'
        ws[f"E{row}"].alignment     = Alignment(horizontal="right", vertical="center")

    ws.row_dimensions[row].height = 26

    # ── NOTAS ───────────────────────────────────────────────────
    if notas:
        row += 2
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = f"Notas: {notas}"
        ws[f"A{row}"].font      = Font(italic=True, size=9, color="666666")
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30

    # ── PIE ─────────────────────────────────────────────────────
    row += 2
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = f"Presupuesto generado el {fecha}  ·  Válido por 48 hs."
    ws[f"A{row}"].font      = Font(size=8, color="999999")
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

    # ── GUARDAR ─────────────────────────────────────────────────
    if destino is None:
        carpeta = Path(__file__).parent.parent / "Presupuestos"
        carpeta.mkdir(parents=True, exist_ok=True)
        destino = carpeta / f"Presupuesto_{numero:05d}_{re.sub(r'[^\\w]','_',cliente)}.xlsx"

    wb.save(destino)
    return destino


def _thin_border(light: bool = False):
    color = "CCCCCC" if light else "AAAAAA"
    side  = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)
